"""Format detector.

Given any source string — local file, folder, http(s) URL, or cloud path
(s3://, gs://, gcs://, az://, abfss://) — determine the storage scheme, the
logical format, compression, whether it is a folder/partitioned dataset, and
(for delimited text) the delimiter, header presence, and encoding.

Detection is layered: cheap signals first (scheme, extension), then a bounded
content sniff for local delimited/text files. Nothing is loaded fully into
memory — the sniff reads at most a few KB.
"""

from __future__ import annotations

import csv
import gzip
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from ai_data_platform.core.exceptions import FormatDetectionError
from ai_data_platform.core.logging import get_logger

log = get_logger("adp.ingestion.detector")

# Logical formats the engine understands.
FORMATS = (
    "csv", "tsv", "json", "ndjson", "parquet", "excel", "arrow", "orc",
    "avro", "sqlite", "delta", "iceberg", "postgres_sql",
)

_CLOUD_SCHEMES = ("s3", "gs", "gcs", "az", "azure", "abfs", "abfss", "http", "https")

_EXT_FORMAT: dict[str, str] = {
    ".csv": "csv",
    ".tsv": "tsv",
    ".txt": "csv",
    ".json": "json",
    ".ndjson": "ndjson",
    ".jsonl": "ndjson",
    ".parquet": "parquet",
    ".pq": "parquet",
    ".xlsx": "excel",
    ".xlsm": "excel",
    ".xls": "excel",
    ".arrow": "arrow",
    ".feather": "arrow",
    ".ipc": "arrow",
    ".orc": "orc",
    ".avro": "avro",
    ".sqlite": "sqlite",
    ".sqlite3": "sqlite",
    ".db": "sqlite",
    ".duckdb": "sqlite",  # treated via sqlite/duckdb attach path
    ".sql": "postgres_sql",
}

_COMPRESSION_EXT = {".gz": "gzip", ".zip": "zip", ".zst": "zstd", ".snappy": "snappy", ".bz2": "bzip2"}

# Cap on files inspected when guessing a folder's dominant format (bounds the walk).
_FOLDER_SCAN_LIMIT = 1000


@dataclass
class Detection:
    source_path: str
    scheme: str  # local | s3 | gs | az | http
    fmt: str
    compression: str | None = None
    is_folder: bool = False
    is_glob: bool = False
    partitioned: bool = False
    partition_keys: list[str] = field(default_factory=list)
    delimiter: str | None = None
    has_header: bool | None = None
    encoding: str | None = None
    excel_sheets: list[str] = field(default_factory=list)
    notes: list[str] = field(default_factory=list)
    extras: dict[str, Any] = field(default_factory=dict)


def _scheme(path: str) -> str:
    m = re.match(r"^([a-zA-Z][a-zA-Z0-9+.-]*)://", path)
    if not m:
        return "local"
    s = m.group(1).lower()
    if s in ("s3", "s3a"):
        return "s3"
    if s in ("gs", "gcs"):
        return "gs"
    if s in ("az", "azure", "abfs", "abfss"):
        return "az"
    if s in ("http", "https"):
        return "http"
    return s


def _strip_compression(name: str) -> tuple[str, str | None]:
    p = Path(name)
    comp = _COMPRESSION_EXT.get(p.suffix.lower())
    if comp:
        return p.stem, comp
    return name, None


def _format_from_name(name: str) -> str | None:
    base, _ = _strip_compression(name)
    return _EXT_FORMAT.get(Path(base).suffix.lower())


def detect(source_path: str, options: dict[str, Any] | None = None) -> Detection:
    """Detect scheme, format, compression, and structure for a source."""
    options = options or {}
    scheme = _scheme(source_path)
    d = Detection(source_path=source_path, scheme=scheme, fmt="")

    # Explicit override wins.
    if options.get("format"):
        d.fmt = str(options["format"]).lower()
    d.is_glob = "*" in source_path or "?" in source_path

    if scheme == "local" and not d.is_glob:
        _detect_local(source_path, d, options)
    else:
        # Remote or glob: rely on extension/override; content sniff not attempted.
        base_name = source_path.rstrip("/").split("/")[-1]
        _, d.compression = _strip_compression(base_name)
        if not d.fmt:
            d.fmt = _format_from_name(base_name) or _format_from_name(source_path) or ""
        if source_path.rstrip("/").split("/")[-1] == "" or d.is_glob:
            d.is_folder = not d.is_glob
        d.notes.append(f"remote/glob source ({scheme}); relying on extension/override")

    if not d.fmt:
        raise FormatDetectionError(
            f"Could not determine the format of {source_path!r}.",
            hint="Pass options={'format': 'csv'|'parquet'|…} to force it.",
        )
    if d.fmt not in FORMATS:
        raise FormatDetectionError(f"Unknown format {d.fmt!r} for {source_path!r}.")
    log.info("detected format=%s scheme=%s compression=%s folder=%s",
             d.fmt, d.scheme, d.compression, d.is_folder)
    return d


def _detect_local(source_path: str, d: Detection, options: dict[str, Any]) -> None:
    p = Path(source_path).expanduser()
    if not p.exists():
        raise FormatDetectionError(
            f"Path {p} does not exist.", hint="Check the path or pass a glob/URL."
        )
    if p.is_dir():
        d.is_folder = True
        _detect_folder(p, d, options)
        return

    _, d.compression = _strip_compression(p.name)
    if not d.fmt:
        d.fmt = _format_from_name(p.name) or ""

    # Content sniff for delimited/uncompressed text (or gzip) when format is
    # text-like or still unknown.
    if d.fmt in ("csv", "tsv", "") and d.compression in (None, "gzip"):
        _sniff_delimited(p, d, options)


def _detect_folder(folder: Path, d: Detection, options: dict[str, Any]) -> None:
    # Partitioned dataset? hive-style key=value subdirs. Only the FIRST parquet
    # path is needed to confirm the layout — avoid listing the whole tree, which
    # can be 10^5+ files on a real data lake.
    hive = re.compile(r"[^/=]+=[^/]+")
    first_parquet = next(folder.rglob("*.parquet"), None)
    if first_parquet is not None:
        d.fmt = d.fmt or "parquet"
        rel = first_parquet.relative_to(folder)
        keys = [seg.split("=")[0] for seg in rel.parts if hive.fullmatch(seg)]
        if keys:
            d.partitioned = True
            d.partition_keys = keys
        return
    # Otherwise pick the dominant file format, sampling at most a bounded number
    # of files (enough to decide the majority without walking a huge tree).
    counts: dict[str, int] = {}
    seen = 0
    for f in folder.rglob("*"):
        if not f.is_file():
            continue
        fmt = _format_from_name(f.name)
        if fmt:
            counts[fmt] = counts.get(fmt, 0) + 1
            seen += 1
            if seen >= _FOLDER_SCAN_LIMIT:
                break
    if not counts:
        raise FormatDetectionError(
            f"No recognized data files under folder {folder}.",
            hint="Supported extensions include csv, json, parquet, orc, avro, arrow.",
        )
    d.fmt = d.fmt or max(counts, key=lambda k: counts[k])
    d.notes.append(f"folder with formats {counts}; using {d.fmt}")


def _sniff_delimited(path: Path, d: Detection, options: dict[str, Any]) -> None:
    encoding = options.get("encoding") or _sniff_encoding(path, d.compression)
    d.encoding = encoding
    opener = gzip.open if d.compression == "gzip" else open
    try:
        with opener(path, "rt", encoding=encoding, errors="replace") as fh:  # type: ignore[operator]
            sample = fh.read(16384)
    except OSError as e:  # pragma: no cover - unreadable file
        raise FormatDetectionError(f"Cannot read {path}: {e}") from e
    if not sample.strip():
        d.fmt = d.fmt or "csv"
        d.notes.append("empty or whitespace-only file")
        return

    delim = options.get("delimiter")
    header = options.get("has_header")
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",\t;|")
        delim = delim or dialect.delimiter
    except csv.Error:
        delim = delim or ("\t" if d.fmt == "tsv" else ",")
        d.notes.append("delimiter sniff inconclusive; used default")
    if header is None:
        try:
            header = csv.Sniffer().has_header(sample)
        except csv.Error:
            header = True
    d.delimiter = delim
    d.has_header = bool(header)
    if not d.fmt:
        d.fmt = "tsv" if delim == "\t" else "csv"


def _sniff_encoding(path: Path, compression: str | None) -> str:
    raw = b""
    try:
        if compression == "gzip":
            with gzip.open(path, "rb") as fh:
                raw = fh.read(8192)
        else:
            with open(path, "rb") as fh:
                raw = fh.read(8192)
    except OSError:
        return "utf-8"
    # Optional dependency: charset-normalizer gives a better guess when present.
    try:
        from charset_normalizer import from_bytes

        best = from_bytes(raw).best()
        if best and best.encoding:
            return best.encoding
    except Exception:  # noqa: BLE001 - optional; fall back to heuristic
        pass
    try:
        raw.decode("utf-8")
        return "utf-8"
    except UnicodeDecodeError:
        return "latin-1"


def excel_sheet_names(path: str) -> list[str]:
    """List sheet names in an Excel workbook (openpyxl for xlsx, pandas for xls)."""
    p = Path(path).expanduser()
    if p.suffix.lower() in (".xlsx", ".xlsm"):
        try:
            from openpyxl import load_workbook

            wb = load_workbook(p, read_only=True)
            try:
                return list(wb.sheetnames)
            finally:
                wb.close()
        except Exception as e:  # noqa: BLE001
            raise FormatDetectionError(f"Cannot read Excel workbook {p}: {e}") from e
    try:
        import pandas as pd

        return list(pd.ExcelFile(p).sheet_names)
    except Exception as e:  # noqa: BLE001
        raise FormatDetectionError(f"Cannot read Excel workbook {p}: {e}") from e
